"""Download a Google Sheet as XLSX and enumerate its tabs.

The Drive `download_file(file_id, export_mime=XLSX)` path preserves all
tabs, unlike the CSV export which only returns the first one.
"""

from __future__ import annotations

import io
from typing import Optional

import pandas as pd
from openpyxl import load_workbook

from api.integrations.google_drive import DriveClient, get_drive_client


XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
GOOGLE_SHEET_MIME = "application/vnd.google-apps.spreadsheet"


def _get_file_mime(file_id: str, *, client: DriveClient) -> str:
    """Drive doesn't expose mimeType on the download client, so go direct."""
    resp = (
        client._service.files()
        .get(fileId=file_id, fields="mimeType", supportsAllDrives=True)
        .execute()
    )
    return resp.get("mimeType", "")


def download_google_sheet_as_xlsx(
    file_id: str, *, client: Optional[DriveClient] = None
) -> bytes:
    """Return XLSX bytes for a spreadsheet file by id.

    Native Google Sheets get exported to XLSX. Files that are already XLSX
    (uploaded .xlsx in Drive) are downloaded as-is.
    """
    client = client or get_drive_client()
    mime = _get_file_mime(file_id, client=client)
    if mime == GOOGLE_SHEET_MIME:
        return client.download_file(file_id, export_mime=XLSX_MIME)
    if mime == XLSX_MIME:
        return client.download_file(file_id)
    raise ValueError(
        f"Drive file {file_id} has mime {mime!r}; expected a Google Sheet or "
        f"an uploaded XLSX. Convert the file or point to a different one."
    )


def list_tab_names(xlsx_bytes: bytes) -> list[str]:
    wb = load_workbook(io.BytesIO(xlsx_bytes), read_only=True, data_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def read_tab(xlsx_bytes: bytes, tab_name: str, *, header_row: int = 0) -> pd.DataFrame:
    """Read one tab into a string-typed DataFrame. `header_row` is 0-indexed."""
    return pd.read_excel(
        io.BytesIO(xlsx_bytes),
        sheet_name=tab_name,
        dtype=str,
        engine="openpyxl",
        header=header_row,
        keep_default_na=False,
        na_values=[""],
    )


def read_tab_no_header(xlsx_bytes: bytes, tab_name: str) -> pd.DataFrame:
    """Read one tab without treating any row as a header. Useful for sheets
    that mix multiple tables in one tab."""
    return pd.read_excel(
        io.BytesIO(xlsx_bytes),
        sheet_name=tab_name,
        dtype=str,
        engine="openpyxl",
        header=None,
        keep_default_na=False,
        na_values=[""],
    )
