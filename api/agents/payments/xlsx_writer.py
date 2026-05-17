"""Zoho payment-upload schedule XLSX writer.

Generates a workbook with one row per allocation (one invoice = one row).
Columns are aligned to Zoho Books / Zoho Billing's "Import Customer Payments"
template so Finance can drop it straight into Zoho.

Reference columns Zoho expects (case-insensitive on import):
  Customer Name, Invoice Number, Invoice Amount, Amount, Payment Date,
  Payment Mode, Reference#, Description
"""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .engine import ReconcileResult


_PAYMENT_MODE_MAP = {
    "mtn": "MTN MoMo",
    "telecel": "Telecel Cash",
    "hero": "Hero",
    "bank": "Bank Transfer",
    "cash": "Cash",
    "bolt_deduction": "Bolt Deduction",
    "unknown": "Other",
}


def _money(d: Decimal) -> float:
    return float(round(d, 2))


def write_zoho_schedule(result: ReconcileResult) -> bytes:
    wb = Workbook()
    # ----- Sheet 1: Zoho upload (allocations) -----
    ws = wb.active
    ws.title = "Zoho upload"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="16222D")
    border = Border(*(Side(border_style="thin", color="E8E2D6"),) * 4)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    headers = [
        "Customer Name",
        "Customer ID",
        "Invoice Number",
        "Payment Date",
        "Amount",
        "Payment Mode",
        "Reference#",
        "Description",
        "Source File",
        "Match Method",
        "Match Confidence",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    row = 2
    for rp in result.matched:
        pay_date = rp.payment.date.isoformat() if rp.payment.date else ""
        mode = _PAYMENT_MODE_MAP.get(rp.payment.channel, "Other")
        for a in rp.allocations:
            ws.cell(row=row, column=1, value=rp.rider_name).alignment = left
            ws.cell(row=row, column=2, value=rp.rider_id).alignment = left
            ws.cell(row=row, column=3, value=a.invoice_number).alignment = left
            ws.cell(row=row, column=4, value=pay_date).alignment = center
            cell_amount = ws.cell(row=row, column=5, value=_money(a.applied_ghs))
            cell_amount.number_format = "#,##0.00"
            cell_amount.alignment = center
            ws.cell(row=row, column=6, value=mode).alignment = center
            ws.cell(row=row, column=7, value=rp.payment.reference).alignment = left
            desc = f"Reconciled from {rp.payment.source_file} line {rp.payment.line_no}"
            ws.cell(row=row, column=8, value=desc).alignment = left
            ws.cell(row=row, column=9, value=rp.payment.source_file).alignment = left
            ws.cell(row=row, column=10, value=rp.method).alignment = center
            ws.cell(row=row, column=11, value=f"{rp.confidence:.0%}").alignment = center
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = border
            row += 1

    # Column widths
    widths = [28, 14, 18, 13, 12, 16, 22, 38, 30, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # ----- Sheet 2: Unmatched (to suspense) -----
    ws2 = wb.create_sheet("Unmatched (to suspense)")
    headers2 = [
        "Source File",
        "Line",
        "Payment Date",
        "Amount",
        "Channel",
        "Sender / Narration",
        "MSISDN",
        "Reference",
        "Best-guess Rider",
        "Confidence",
        "Reason",
    ]
    for i, h in enumerate(headers2, start=1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = border

    for r_idx, u in enumerate(result.unmatched, start=2):
        p = u.payment
        ws2.cell(row=r_idx, column=1, value=p.source_file).alignment = left
        ws2.cell(row=r_idx, column=2, value=p.line_no).alignment = center
        ws2.cell(row=r_idx, column=3, value=p.date.isoformat() if p.date else "").alignment = center
        amt = ws2.cell(row=r_idx, column=4, value=_money(p.amount_ghs))
        amt.number_format = "#,##0.00"
        amt.alignment = center
        ws2.cell(row=r_idx, column=5, value=p.channel).alignment = center
        ws2.cell(row=r_idx, column=6, value=p.raw_name).alignment = left
        ws2.cell(row=r_idx, column=7, value=p.msisdn or "").alignment = left
        ws2.cell(row=r_idx, column=8, value=p.reference).alignment = left
        ws2.cell(row=r_idx, column=9, value=u.best_guess_rider_name).alignment = left
        ws2.cell(row=r_idx, column=10, value=f"{u.best_guess_confidence:.0%}").alignment = center
        ws2.cell(row=r_idx, column=11, value=u.reason).alignment = left
        for col in range(1, len(headers2) + 1):
            ws2.cell(row=r_idx, column=col).border = border

    widths2 = [30, 8, 13, 12, 14, 32, 14, 22, 24, 12, 36]
    for i, w in enumerate(widths2, start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # ----- Sheet 3: Summary -----
    ws3 = wb.create_sheet("Summary")
    ws3.column_dimensions["A"].width = 36
    ws3.column_dimensions["B"].width = 28
    title = ws3.cell(row=1, column=1, value="Payment reconciliation summary")
    title.font = Font(size=14, bold=True)
    pairs = [
        ("Cutoff date (in-scope from)", result.cutoff_date.isoformat()),
        ("Invoices in corpus", result.invoices_corpus_size),
        ("Riders in master", result.riders_in_master),
        ("Total payments parsed", result.total_payments),
        ("Payments in scope (date >= cutoff)", result.in_scope_payments),
        ("Matched payments", len(result.matched)),
        ("Unmatched payments → suspense", len(result.unmatched)),
        ("Total matched amount (GHS)", _money(result.total_matched_amount_ghs)),
        ("Total unmatched amount (GHS)", _money(result.total_unmatched_amount_ghs)),
    ]
    for i, (k, v) in enumerate(pairs, start=3):
        ws3.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws3.cell(row=i, column=2, value=v)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
