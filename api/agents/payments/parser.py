"""Flexible payment statement parser.

Handles a wide range of MoMo / bank / Wahu-internal CSV and XLSX shapes by
fuzzy-matching column headers. Each output row is a normalised PaymentRow.
"""

from __future__ import annotations

import csv
import logging
import re
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable, Optional

logger = logging.getLogger(__name__)


@dataclass
class PaymentRow:
    source_file: str
    line_no: int
    date: Optional[date]
    amount_ghs: Decimal
    raw_name: str            # sender / "From name"
    msisdn: Optional[str]    # phone number if found
    reference: str           # MoMo External id / bank Reference / Cheque
    channel: str             # mtn | telecel | hero | bank | bolt_deduction | cash | unknown
    narration: str = ""      # freeform message ("To message", "Particulars")
    raw: dict = None         # original row, for debugging / audit

    def __post_init__(self):
        if self.raw is None:
            self.raw = {}


# Column aliases — case-insensitive substring match.
_DATE_KEYS = ("date", "txn date", "transaction date", "value date", "posted")
_AMOUNT_KEYS = ("amount", "credit", "credit amount", "amount paid", "amount (ghs)", "total")
_NAME_KEYS = ("from name", "sender name", "name", "narration", "description", "details", "customer", "sender", "payer")
_PHONE_KEYS = ("from account", "phone", "msisdn", "mobile", "number", "sender phone", "msisdn / phone")
# Order matters: MoMo statements have BOTH "Id" (internal wallet record)
# and "External id" (the actual MoMo Txn ID Finance recognises). Bank
# statements use "Reference" / "Cheque". We list the specific MoMo and
# bank aliases first so they win over generic shorter strings.
_REF_KEYS = (
    "external id", "external ref", "external reference",
    "transaction id", "transaction ref", "reference number",
    "receipt number", "receipt no", "receipt",
    "txn id", "trxn id", "trans id",
    "wallet id", "voucher", "cheque",
    "reference", "ref",
)
# Narration = the freeform message the sender attached to the payment
# (MoMo "To message", bank statement "Particulars"/"Description"). Kept
# distinct from sender name so the UI can show both.
_NARRATION_KEYS = (
    "to message", "message", "narration", "remarks", "particulars",
    "memo", "description",
)
_CHANNEL_KEYS = ("channel", "mode", "payment mode", "method", "source", "type", "provider category")


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _find_col(headers: list[str], aliases: tuple[str, ...]) -> Optional[str]:
    cols = _find_all_cols(headers, aliases)
    return cols[0] if cols else None


def _find_all_cols(headers: list[str], aliases: tuple[str, ...]) -> list[str]:
    """Return ALL headers matching any alias, ordered by alias priority.

    Used as a fallback chain at row-parse time: e.g. for the Reference
    field we try "External id" first, then fall back to "Id" for rows
    where the External id is blank (internal MoMo transfers etc.)."""
    norm_headers = [(h, _norm(h)) for h in headers]
    seen: set[str] = set()
    out: list[str] = []
    # Exact phase: walk aliases first so the most-specific alias wins.
    for a in aliases:
        for original, n in norm_headers:
            if n == a and original not in seen:
                out.append(original)
                seen.add(original)
    # Substring phase: same order — earlier alias = higher priority.
    for a in aliases:
        for original, n in norm_headers:
            if a in n and original not in seen:
                out.append(original)
                seen.add(original)
    return out


def _first_nonempty(row: dict, cols: list[str]) -> str:
    for c in cols:
        v = str(row.get(c, "") or "").strip()
        if v:
            return v
    return ""


def _parse_date(v: str) -> Optional[date]:
    v = (v or "").strip()
    if not v:
        return None
    fmts = (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%d %b %Y",
        "%d-%b-%Y",
        "%b %d, %Y",
        "%m/%d/%Y",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y %H:%M:%S",
    )
    for f in fmts:
        try:
            return datetime.strptime(v[: len(f) + 4], f).date()
        except ValueError:
            continue
    # Last resort: ISO-ish prefix
    try:
        return datetime.fromisoformat(v.split("T")[0]).date()
    except Exception:
        return None


def _parse_amount(v: str) -> Optional[Decimal]:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    # Strip currency symbols / commas / parens (parens = negative)
    neg = s.startswith("(") and s.endswith(")")
    s = re.sub(r"[^\d\.\-]", "", s)
    if not s:
        return None
    try:
        d = Decimal(s)
    except InvalidOperation:
        return None
    if neg:
        d = -d
    return d


_GHANA_PHONE_RE = re.compile(r"(?:\+?233|0)(2\d{8}|5\d{8})")


def _extract_phone(*parts: str) -> Optional[str]:
    for p in parts:
        if not p:
            continue
        m = _GHANA_PHONE_RE.search(str(p))
        if m:
            digits = m.group(1)
            return "0" + digits  # canonical 10-digit Ghanaian form
    return None


def _detect_channel(channel_text: str, name_text: str) -> str:
    blob = f"{channel_text} {name_text}".lower()
    if "mtn" in blob or "momo" in blob or "mobile money" in blob:
        return "mtn"
    if "telecel" in blob or "vodafone" in blob:
        return "telecel"
    if "hero" in blob or "airtel" in blob:
        return "hero"
    if "bolt" in blob:
        return "bolt_deduction"
    if any(t in blob for t in ("bank", "transfer", "absa", "stanbic", "fidelity", "gcb", "ecobank", "cal", "uba")):
        return "bank"
    if "cash" in blob:
        return "cash"
    return "unknown"


def _iter_csv(path: Path) -> Iterable[dict]:
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            yield row


def _iter_xlsx(path: Path) -> Iterable[dict]:
    # Lazy import — only loaded if an xlsx is actually present.
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if ws is None:
        return
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(h or "").strip() for h in next(rows)]
    except StopIteration:
        return
    for r in rows:
        if all(v is None or v == "" for v in r):
            continue
        yield {headers[i]: ("" if v is None else v) for i, v in enumerate(r) if i < len(headers)}


def parse_payment_file(path: str | Path) -> list[PaymentRow]:
    """Parse one payment file (CSV or XLSX). Returns a list of normalised rows."""
    p = Path(path)
    if not p.exists():
        return []

    if p.suffix.lower() in (".xlsx", ".xls"):
        iterator = _iter_xlsx(p)
    else:
        iterator = _iter_csv(p)

    first = next(iter(iterator), None)
    if not first:
        return []

    headers = list(first.keys())
    col_date = _find_col(headers, _DATE_KEYS)
    col_amount = _find_col(headers, _AMOUNT_KEYS)
    # Use chains so per-row fallback works when the preferred column is
    # blank for a given transaction (common with MoMo's optional fields).
    name_chain = _find_all_cols(headers, _NAME_KEYS)
    phone_chain = _find_all_cols(headers, _PHONE_KEYS)
    ref_chain = _find_all_cols(headers, _REF_KEYS)
    narration_chain = _find_all_cols(headers, _NARRATION_KEYS)
    col_channel = _find_col(headers, _CHANNEL_KEYS)

    if not col_amount:
        logger.warning("payment parser: %s has no amount column; headers=%s", p.name, headers)
        return []

    # Re-stream from the top: combine the first row with the iterator
    def streamer():
        yield first
        for r in iterator:
            yield r

    out: list[PaymentRow] = []
    for i, row in enumerate(streamer(), start=2):  # row 1 = header
        amount = _parse_amount(str(row.get(col_amount, "")))
        if amount is None or amount <= 0:
            continue
        raw_name = _first_nonempty(row, name_chain)
        phone_blob = _first_nonempty(row, phone_chain)
        msisdn = _extract_phone(phone_blob, raw_name, _first_nonempty(row, ref_chain))
        # If extract_phone couldn't validate a Ghana number but we DO have a
        # raw "From account" value, keep it as the displayed phone so the
        # UI shows something instead of "—".
        if not msisdn and phone_blob:
            msisdn = phone_blob
        ref = _first_nonempty(row, ref_chain)
        narration = _first_nonempty(row, narration_chain)
        channel = _detect_channel(
            str(row.get(col_channel, "")) if col_channel else "",
            raw_name,
        )
        out.append(
            PaymentRow(
                source_file=p.name,
                line_no=i,
                date=_parse_date(str(row.get(col_date, ""))) if col_date else None,
                amount_ghs=amount,
                raw_name=raw_name,
                msisdn=msisdn,
                reference=ref,
                channel=channel,
                narration=narration,
                raw=dict(row),
            )
        )
    return out


def parse_folder(folder: str | Path) -> list[PaymentRow]:
    """Parse every CSV/XLSX in a folder."""
    p = Path(folder)
    if not p.exists():
        return []
    out: list[PaymentRow] = []
    for child in sorted(p.iterdir()):
        if child.is_file() and child.suffix.lower() in (".csv", ".xlsx", ".xls"):
            try:
                out.extend(parse_payment_file(child))
            except Exception as e:
                logger.exception("failed to parse %s: %s", child.name, e)
    return out
