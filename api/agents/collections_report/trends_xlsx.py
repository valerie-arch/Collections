"""Render TrendsReport → multi-sheet xlsx for the Portfolio Trends download."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from api.agents.collections_report.trends import TrendsReport, RiderRanking

INK = "0F1419"
ACCENT = "B8763E"
LINE = "E8E2D6"

H1 = Font(name="Calibri", size=16, bold=True, color=INK)
H2 = Font(name="Calibri", size=12, bold=True, color=INK)
SUB = Font(name="Calibri", size=10, italic=True, color="4A5159")
HEAD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor=INK)

THIN = Side(style="thin", color=LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY = "#,##0.00"
PCT = "0.00%"
INT = "#,##0"


def _head(ws, row: int, headers: list[str]) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEAD
        c.fill = HEAD_FILL


def _write_summary(ws, r: TrendsReport, fleet: str) -> None:
    ws.title = "Summary"
    ws["A1"] = f"Wahu Portfolio Trends — {fleet}"
    ws["A1"].font = H1
    ws["A2"] = (
        f"As-of {r.as_of.isoformat()} · {len(r.months)} months covered "
        f"({r.months[0].label} → {r.months[-1].label})"
    )
    ws["A2"].font = SUB

    ws["A4"] = "Cumulative status"
    ws["A4"].font = H2

    _head(ws, 5, ["Status", "Riders"])
    rows = [
        ("Active subscriptions", r.cumulative_active),
        ("In recovery (churned w/ debt)", r.cumulative_recovery),
        ("Completed (fully paid out)", r.cumulative_completed),
        ("Total", r.cumulative_active + r.cumulative_recovery + r.cumulative_completed),
    ]
    for i, (lbl, val) in enumerate(rows, start=6):
        ws.cell(row=i, column=1, value=lbl).border = BORDER
        c = ws.cell(row=i, column=2, value=val)
        c.border = BORDER
        c.number_format = INT
    ws.cell(row=6 + len(rows) - 1, column=1).font = H2
    ws.cell(row=6 + len(rows) - 1, column=2).font = H2

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 14


def _write_monthly(ws, r: TrendsReport) -> None:
    ws.title = "Monthly trends"
    ws["A1"] = "Month-over-month"
    ws["A1"].font = H1

    headers = [
        "Month", "Year", "Active riders", "New riders",
        "Invoices issued", "Invoiced (GHS)", "Collected (GHS)",
        "Collection ratio", "Outstanding at end (GHS)", "MRR (GHS)",
    ]
    _head(ws, 3, headers)

    for i, m in enumerate(r.months, start=4):
        ratio = (float(m.collected_ghs) / float(m.invoiced_ghs)) if m.invoiced_ghs > 0 else 0.0
        vals = [
            m.label,
            m.year,
            m.active_riders,
            m.new_riders,
            m.invoices_issued,
            float(m.invoiced_ghs),
            float(m.collected_ghs),
            ratio,
            float(m.outstanding_ghs),
            float(m.mrr_ghs),
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = BORDER
            if col in (3, 4, 5):
                cell.number_format = INT
            elif col in (6, 7, 9, 10):
                cell.number_format = MONEY
            elif col == 8:
                cell.number_format = PCT

    widths = (10, 8, 14, 12, 16, 18, 18, 16, 22, 14)
    for col, w in zip(range(1, len(widths) + 1), widths):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A4"


def _rank_sheet(ws, title: str, riders: list[RiderRanking]) -> None:
    ws.title = title[:31]
    ws["A1"] = title
    ws["A1"].font = H1
    headers = [
        "Customer", "Customer ID", "Invoiced (GHS)",
        "Collected (GHS)", "Outstanding (GHS)", "Collection ratio",
    ]
    _head(ws, 3, headers)
    for i, rd in enumerate(riders, start=4):
        vals = [
            rd.customer_name,
            rd.customer_id,
            float(rd.lifetime_invoiced_ghs),
            float(rd.lifetime_collected_ghs),
            float(rd.lifetime_outstanding_ghs),
            rd.collection_ratio,
        ]
        for col, v in enumerate(vals, start=1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.border = BORDER
            if col in (3, 4, 5):
                cell.number_format = MONEY
            elif col == 6:
                cell.number_format = PCT
    widths = (32, 16, 18, 18, 20, 16)
    for col, w in zip(range(1, len(widths) + 1), widths):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A4"


def _write_methodology(ws, fleet: str) -> None:
    ws.title = "Methodology"
    ws["A1"] = "How this report is built"
    ws["A1"].font = H1
    rows = [
        ("Topic", "Note"),
        ("Source", "Zoho Billing invoice CSV exports, deduped across files by Invoice ID."),
        ("Fleet split",
         "Wahu OS Bikes_and_Riders.xlsx provides the authoritative rider → fleet "
         "mapping by matching Assigned Rider name. Zoho subscription TSA flag is "
         "secondary. Default Wahu."),
        ("Fleet filter applied", fleet),
        ("Active subscriptions",
         "Customers whose Zoho subscription is currently live or paused."),
        ("In recovery",
         "Customers cancelled in Zoho with outstanding balance."),
        ("Completed",
         "Customers whose Zoho subscription has expired (fully paid out)."),
        ("MRR proxy",
         "Total invoiced in the calendar month — treated as MRR since billing "
         "is mostly monthly. For weekly subs this overstates slightly."),
        ("Collected (monthly)",
         "Sum of (invoice total − balance) for invoices whose Last Payment Date "
         "falls in the month. Older invoice exports may underreport recent "
         "payments; re-export older periods from Zoho for accuracy."),
        ("Outstanding at month-end",
         "Sum of balances on invoices dated in the month (snapshot)."),
        ("Top/Bottom rankings",
         "Rider lifetime totals across all invoices in the dataset. Bottom-10 is "
         "filtered to riders with at least GHS 500 lifetime invoiced to avoid noise."),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        if i == 3:
            ws.cell(row=i, column=1).font = HEAD
            ws.cell(row=i, column=2).font = HEAD
            ws.cell(row=i, column=1).fill = HEAD_FILL
            ws.cell(row=i, column=2).fill = HEAD_FILL
        else:
            ws.cell(row=i, column=1).border = BORDER
            ws.cell(row=i, column=2).border = BORDER
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 90


def write_trends_xlsx(report: TrendsReport, *, fleet: str = "All") -> bytes:
    wb = Workbook()
    _write_summary(wb.active, report, fleet)
    _write_monthly(wb.create_sheet(), report)
    _rank_sheet(wb.create_sheet(), "Top 10 outstanding", report.top_10_outstanding)
    _rank_sheet(wb.create_sheet(), "Top 10 lifetime collected", report.top_10_collected_lifetime)
    _rank_sheet(wb.create_sheet(), "Bottom 10 collection ratio", report.bottom_10_ratio)
    _write_methodology(wb.create_sheet(), fleet)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
