"""Render daily activity log → xlsx for email + Drive upload."""

from __future__ import annotations

from collections import Counter
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


INK = "0F1419"
LINE = "E8E2D6"
H1 = Font(name="Calibri", size=16, bold=True, color=INK)
SUB = Font(name="Calibri", size=10, italic=True, color="4A5159")
HEAD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor=INK)
THIN = Side(style="thin", color=LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_activities_xlsx(day: date, activities: list[dict]) -> bytes:
    riders = {a.get("customer_id") for a in activities if a.get("customer_id")}
    by_action = Counter(a.get("action", "other") for a in activities)
    by_agency = Counter((a.get("agency") or "Unassigned") for a in activities)

    wb = Workbook()

    # ---- Summary sheet ----
    s = wb.active
    s.title = "Summary"
    s["A1"] = f"Wahu Collections — Daily Activities {day.isoformat()}"
    s["A1"].font = H1
    s["A2"] = (
        f"{len(activities)} actions logged · "
        f"{len(riders)} unique riders engaged"
    )
    s["A2"].font = SUB

    s["A4"] = "Actions taken"
    s["A4"].font = HEAD
    s["A4"].fill = HEAD_FILL
    s["B4"] = "Count"
    s["B4"].font = HEAD
    s["B4"].fill = HEAD_FILL
    for i, (action, count) in enumerate(sorted(by_action.items()), start=5):
        s.cell(row=i, column=1, value=action).border = BORDER
        s.cell(row=i, column=2, value=count).border = BORDER

    row = 5 + len(by_action) + 1
    s.cell(row=row, column=1, value="Agency split").font = HEAD
    s.cell(row=row, column=1).fill = HEAD_FILL
    s.cell(row=row, column=2, value="Count").font = HEAD
    s.cell(row=row, column=2).fill = HEAD_FILL
    for i, (agency, count) in enumerate(sorted(by_agency.items()), start=row + 1):
        s.cell(row=i, column=1, value=agency).border = BORDER
        s.cell(row=i, column=2, value=count).border = BORDER

    s.column_dimensions["A"].width = 32
    s.column_dimensions["B"].width = 14

    # ---- Detail sheet ----
    d = wb.create_sheet("Activities")
    d["A1"] = f"Activity log — {day.isoformat()}"
    d["A1"].font = H1

    headers = ["Time (UTC)", "Customer", "Customer ID", "Action", "Agency", "Actor", "Note"]
    for i, h in enumerate(headers, start=1):
        c = d.cell(row=3, column=i, value=h)
        c.font = HEAD
        c.fill = HEAD_FILL

    rows = sorted(activities, key=lambda a: a.get("created_at", ""))
    for r, a in enumerate(rows, start=4):
        d.cell(row=r, column=1, value=(a.get("created_at") or "").replace("T", " ").replace("Z", ""))
        d.cell(row=r, column=2, value=a.get("customer_name", ""))
        d.cell(row=r, column=3, value=a.get("customer_id", ""))
        d.cell(row=r, column=4, value=a.get("action", ""))
        d.cell(row=r, column=5, value=a.get("agency") or "—")
        d.cell(row=r, column=6, value=a.get("actor", ""))
        d.cell(row=r, column=7, value=a.get("note", ""))
        for col in range(1, 8):
            d.cell(row=r, column=col).border = BORDER

    widths = (20, 28, 22, 24, 12, 22, 80)
    for col, w in zip(range(1, len(widths) + 1), widths):
        d.column_dimensions[get_column_letter(col)].width = w
    d.freeze_panes = "A4"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
