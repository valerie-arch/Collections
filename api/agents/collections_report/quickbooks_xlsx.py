"""Render a QbExport to a QBO-compatible xlsx."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from api.agents.collections_report.quickbooks import QbExport

INK = "0F1419"
LINE = "E8E2D6"

H1 = Font(name="Calibri", size=16, bold=True, color=INK)
SUB = Font(name="Calibri", size=10, italic=True, color="4A5159")
HEAD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
HEAD_FILL = PatternFill("solid", fgColor=INK)
THIN = Side(style="thin", color=LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

MONEY = "#,##0.00"
DATE = "yyyy-mm-dd"


INVOICE_COLS = [
    ("InvoiceNo", 16, str),
    ("Customer", 32, str),
    ("InvoiceDate", 14, DATE),
    ("DueDate", 14, DATE),
    ("Terms", 18, str),
    ("Item", 22, str),
    ("Description", 36, str),
    ("Quantity", 10, "0"),
    ("Rate", 14, MONEY),
    ("Amount", 14, MONEY),
    ("Balance", 14, MONEY),
    ("Status", 14, str),
    ("Class", 12, str),
    ("Currency", 10, str),
]

PAYMENT_COLS = [
    ("Customer", 32, str),
    ("PaymentDate", 14, DATE),
    ("Amount", 14, MONEY),
    ("PaymentMethod", 16, str),
    ("ReferenceNo", 18, str),
    ("AppliedToInvoiceNo", 18, str),
    ("Memo", 60, str),
    ("Class", 12, str),
    ("Currency", 10, str),
]


def _write_header_row(ws, headers: list[str], row: int = 3) -> None:
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HEAD
        c.fill = HEAD_FILL


def _write_data_sheet(ws, export: QbExport) -> None:
    title_label = "Invoices" if export.type == "invoices" else "Payments"
    ws.title = title_label
    ws["A1"] = f"Wahu → QuickBooks {title_label} export"
    ws["A1"].font = H1
    ws["A2"] = (
        f"Period {export.window_start} → {export.window_end} · "
        f"Fleet (Class): {export.fleet} · "
        f"{export.row_count} rows · "
        f"Total GHS {float(export.total_amount):,.2f}"
    )
    ws["A2"].font = SUB

    cols = INVOICE_COLS if export.type == "invoices" else PAYMENT_COLS
    _write_header_row(ws, [c[0] for c in cols], row=3)

    rows = export.invoice_rows if export.type == "invoices" else export.payment_rows
    for r_idx, row in enumerate(rows, start=4):
        if export.type == "invoices":
            values = [
                row.invoice_no, row.customer, row.invoice_date, row.due_date,
                row.terms, row.item, row.description, row.quantity,
                row.rate, row.amount, row.balance, row.status, row.fleet, row.currency,
            ]
        else:
            values = [
                row.customer, row.payment_date, row.amount, row.payment_method,
                row.reference_no, row.applied_to_invoice_no, row.memo,
                row.fleet, row.currency,
            ]
        for c_idx, (col_meta, val) in enumerate(zip(cols, values), start=1):
            _, _, fmt = col_meta
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = BORDER
            if fmt and fmt != str:
                cell.number_format = fmt

    for i, (_, w, _) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A4"


def _write_methodology(ws, export: QbExport) -> None:
    ws.title = "Methodology"
    ws["A1"] = "How to use this file in QuickBooks"
    ws["A1"].font = H1

    rows = [
        ("Topic", "Note"),
        ("Source", "Zoho Billing invoice CSV exports synced from Drive, deduped by Invoice ID."),
        ("Period", f"{export.window_start} → {export.window_end}"),
        ("Fleet (QB Class)",
         "Each row's Class column = Wahu or TSA. Map this to your QBO Class in "
         "Settings → All Lists → Classes before importing for clean P&L by fleet."),
        ("Item", "Defaults to 'Subscription billing'. Edit in QB if your chart of "
                 "accounts uses a different service item name."),
        ("Terms", "Defaults to 'Due on Receipt'. Edit in QB if needed."),
        ("PaymentMethod (payments only)",
         "Defaults to 'MoMo'. Replace per row before import (MTN/Telecel/Cash/etc.)."),
        ("Currency", "GHS. Make sure your QB company is set up to accept GHS."),
        ("Import flow",
         "QuickBooks Online → Settings (gear) → Import data → Invoices (or "
         "Receive Payments) → Browse → upload this xlsx → map columns → Next → Import."),
        ("Skipping duplicates",
         "Re-importing the same InvoiceNo will overwrite (or be rejected, depending "
         "on your import preference). Filter the date range to only new periods."),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
        if i == 3:
            ws.cell(row=i, column=1).font = HEAD
            ws.cell(row=i, column=2).font = HEAD
            ws.cell(row=i, column=1).fill = HEAD_FILL
            ws.cell(row=i, column=2).fill = HEAD_FILL
        else:
            ws.cell(row=i, column=1).border = BORDER
            ws.cell(row=i, column=2).border = BORDER
            ws.cell(row=i, column=2).alignment = ws.cell(row=i, column=2).alignment.copy(wrap_text=True, vertical="top")
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 92


def write_qb_xlsx(export: QbExport) -> bytes:
    wb = Workbook()
    _write_data_sheet(wb.active, export)
    _write_methodology(wb.create_sheet(), export)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
