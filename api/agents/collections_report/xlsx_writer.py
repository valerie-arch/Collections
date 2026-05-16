"""Emit Report A as a 3-sheet .xlsx matching Wahu_Collections_Report (1).xlsx.

Structure: Summary (headlines + bands + ageing), Active Riders (scorecard
table with autofilter), Methodology (static text).
"""

from __future__ import annotations

from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from api.agents.collections_report.engine import ReportData


# --- styling tokens (kept close to the source file) -----------------------

INK = "0F1419"
ACCENT = "B8763E"
LINE = "E8E2D6"

H1 = Font(name="Calibri", size=16, bold=True, color=INK)
H2 = Font(name="Calibri", size=12, bold=True, color=INK)
SUBTITLE = Font(name="Calibri", size=10, italic=True, color="4A5159")
TABLE_HEAD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TABLE_HEAD_FILL = PatternFill("solid", fgColor=INK)
ACCENT_FILL = PatternFill("solid", fgColor=ACCENT)
THIN = Side(style="thin", color=LINE)
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY_FMT = '#,##0.00'
PCT_FMT = '0.00%'
INT_FMT = '#,##0'


def _money(v: Decimal) -> float:
    return float(v)


def _write_summary(ws, r: ReportData) -> None:
    ws.title = "Summary"
    ws["A1"] = "Wahu Mobility — Collections Report (Active Riders)"
    ws["A1"].font = H1
    view_label = "Month-to-date view" if r.view == "mtd" else "Lifetime view"
    ws["A2"] = (
        f"As-of {r.as_of.isoformat()}  ·  "
        f"Window {r.window_start.isoformat()} → {r.window_end.isoformat()}  ·  "
        f"{view_label}"
    )
    ws["A2"].font = SUBTITLE

    ws["A4"] = "Headlines"
    ws["A4"].font = H2

    headlines = [
        ("Metric", "Value"),
        (f"Active riders (last invoice in window)", r.active_riders),
        ("  — % of total loan book",
         (r.active_riders / r.total_rider_population) if r.total_rider_population else 0),
        ("Lifetime invoiced (GHS)", _money(r.lifetime_invoiced_ghs)),
        ("Lifetime collected (GHS)", _money(r.lifetime_collected_ghs)),
        ("Lifetime outstanding (GHS)", _money(r.lifetime_outstanding_ghs)),
        ("Lifetime collection ratio",
         float(r.lifetime_collected_ghs / r.lifetime_invoiced_ghs)
         if r.lifetime_invoiced_ghs > 0 else 0),
        ("Open invoice lines (Zoho)", r.open_invoice_lines),
    ]
    for i, (k, v) in enumerate(headlines, start=5):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
        if i == 5:
            ws.cell(row=i, column=1).font = TABLE_HEAD
            ws.cell(row=i, column=2).font = TABLE_HEAD
            ws.cell(row=i, column=1).fill = TABLE_HEAD_FILL
            ws.cell(row=i, column=2).fill = TABLE_HEAD_FILL
        else:
            ws.cell(row=i, column=1).border = BORDER
            ws.cell(row=i, column=2).border = BORDER
        if isinstance(v, float) and "ratio" in k.lower() or "%" in k:
            ws.cell(row=i, column=2).number_format = PCT_FMT
        elif isinstance(v, (int, float)):
            ws.cell(row=i, column=2).number_format = (
                MONEY_FMT if "GHS" in k else INT_FMT
            )

    # Risk band breakdown
    base = 5 + len(headlines) + 1
    ws.cell(row=base, column=1, value="Risk band breakdown").font = H2
    head_row = base + 1
    for i, h in enumerate(["Band", "Riders", "Outstanding (GHS)", "Definition"], start=1):
        c = ws.cell(row=head_row, column=i, value=h)
        c.font = TABLE_HEAD
        c.fill = TABLE_HEAD_FILL
    total_riders = 0
    total_out = Decimal("0")
    for j, b in enumerate(r.risk_bands, start=head_row + 1):
        ws.cell(row=j, column=1, value=b.band).border = BORDER
        ws.cell(row=j, column=2, value=b.riders).border = BORDER
        ws.cell(row=j, column=3, value=_money(b.outstanding_ghs)).border = BORDER
        ws.cell(row=j, column=3).number_format = MONEY_FMT
        ws.cell(row=j, column=4, value=b.definition).border = BORDER
        total_riders += b.riders
        total_out += b.outstanding_ghs
    total_row = head_row + 1 + len(r.risk_bands)
    ws.cell(row=total_row, column=1, value="Total").font = H2
    ws.cell(row=total_row, column=2, value=total_riders).font = H2
    ws.cell(row=total_row, column=3, value=_money(total_out)).font = H2
    ws.cell(row=total_row, column=3).number_format = MONEY_FMT

    # Ageing profile
    base = total_row + 2
    ws.cell(row=base, column=1, value="⏳ Ageing profile of open invoices").font = H2
    head_row = base + 1
    for i, h in enumerate(["Age bucket", "Open invoices", "Outstanding (GHS)", "% of total"], start=1):
        c = ws.cell(row=head_row, column=i, value=h)
        c.font = TABLE_HEAD
        c.fill = TABLE_HEAD_FILL
    total_open = sum(b.open_invoices for b in r.ageing)
    total_a_out = sum((b.outstanding_ghs for b in r.ageing), Decimal("0"))
    for j, b in enumerate(r.ageing, start=head_row + 1):
        ws.cell(row=j, column=1, value=b.label).border = BORDER
        ws.cell(row=j, column=2, value=b.open_invoices).border = BORDER
        ws.cell(row=j, column=3, value=_money(b.outstanding_ghs)).border = BORDER
        ws.cell(row=j, column=3).number_format = MONEY_FMT
        pct = float(b.outstanding_ghs / total_a_out) if total_a_out > 0 else 0
        ws.cell(row=j, column=4, value=pct).border = BORDER
        ws.cell(row=j, column=4).number_format = PCT_FMT
    last = head_row + 1 + len(r.ageing)
    ws.cell(row=last, column=1, value="Total").font = H2
    ws.cell(row=last, column=2, value=total_open).font = H2
    ws.cell(row=last, column=3, value=_money(total_a_out)).font = H2
    ws.cell(row=last, column=3).number_format = MONEY_FMT
    ws.cell(row=last, column=4, value=1.0 if total_a_out > 0 else 0).font = H2
    ws.cell(row=last, column=4).number_format = PCT_FMT

    for col, width in zip("ABCDE", (38, 18, 22, 28, 14)):
        ws.column_dimensions[col].width = width


def _write_active_riders(ws, r: ReportData) -> None:
    ws.title = "Active Riders"
    ws["A1"] = "Active rider scorecard"
    ws["A1"].font = H1

    headers = [
        "Customer Number", "Customer Name", "First Invoice", "Last Invoice",
        "Months since last invoice", "Lifetime Invoices", "Open Invoices",
        "Lifetime Invoiced GHS", "Lifetime Collected GHS",
        "Lifetime Outstanding GHS", "Collection Ratio", "Risk Band", "Plans",
    ]
    head_row = 3
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=head_row, column=i, value=h)
        c.font = TABLE_HEAD
        c.fill = TABLE_HEAD_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")

    for j, s in enumerate(r.scorecards, start=head_row + 1):
        vals = [
            s.customer_id, s.customer_name,
            s.first_invoice, s.last_invoice,
            s.months_since_last_invoice, s.lifetime_invoices, s.open_invoices,
            _money(s.lifetime_invoiced_ghs), _money(s.lifetime_collected_ghs),
            _money(s.lifetime_outstanding_ghs), s.collection_ratio,
            s.risk_band, s.plans,
        ]
        for i, v in enumerate(vals, start=1):
            cell = ws.cell(row=j, column=i, value=v)
            cell.border = BORDER
            if i in (3, 4):
                cell.number_format = "yyyy-mm-dd"
            elif i in (8, 9, 10):
                cell.number_format = MONEY_FMT
            elif i == 11:
                cell.number_format = PCT_FMT

    last_row = head_row + len(r.scorecards)
    last_col = get_column_letter(len(headers))
    if r.scorecards:
        table = Table(
            displayName="ActiveRiders",
            ref=f"A{head_row}:{last_col}{last_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False,
        )
        ws.add_table(table)

    widths = (16, 30, 13, 13, 12, 12, 12, 18, 18, 20, 14, 10, 28)
    for col, w in zip(range(1, len(widths) + 1), widths):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A4"


def _write_methodology(ws, r: ReportData) -> None:
    ws.title = "Methodology"
    ws["A1"] = "How this report is built"
    ws["A1"].font = H1
    rows = [
        ("Topic", "Note"),
        ("Scope",
         f"Active riders = at least one Zoho invoice in {r.window_start} → {r.window_end} "
         f"({r.active_riders} of {r.total_rider_population} total)."),
        ("Source", "Zoho Billing invoice CSV exports, deduped across files by Invoice ID."),
        ("Lifetime invoiced", "Sum of Total across all invoices for the rider in scope."),
        ("Lifetime outstanding", "Sum of Balance across open invoices."),
        ("Lifetime collected", "Lifetime invoiced − lifetime outstanding."),
        ("Risk band",
         "Per-rider ratio bucketed: A ≥95%, B 80–95%, C 60–80%, D 30–60%, E <30%. "
         "Watch D and E — active riders drifting toward churn."),
        ("Ageing profile",
         "Open invoices (Balance > 0) bucketed by days since invoice date: "
         "0–30, 31–60, 61–90, 91–180, 181–365, 365+. Past 60d slips; 90+ needs intervention."),
        ("View",
         "Toggle Lifetime ↔ MTD on the Reports page. "
         "MTD scopes numbers to invoices in the current month; Lifetime is YTD cohort with all-time totals."),
        ("Generated", f"{r.as_of.isoformat()} by the Collections platform."),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical="top")
        if i == 3:
            ws.cell(row=i, column=1).font = TABLE_HEAD
            ws.cell(row=i, column=2).font = TABLE_HEAD
            ws.cell(row=i, column=1).fill = TABLE_HEAD_FILL
            ws.cell(row=i, column=2).fill = TABLE_HEAD_FILL
        else:
            ws.cell(row=i, column=1).border = BORDER
            ws.cell(row=i, column=2).border = BORDER
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 95


def write_report_xlsx(report: ReportData) -> bytes:
    """Render ReportData → .xlsx bytes."""
    wb = Workbook()
    _write_summary(wb.active, report)
    _write_active_riders(wb.create_sheet("Active Riders"), report)
    _write_methodology(wb.create_sheet("Methodology"), report)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
